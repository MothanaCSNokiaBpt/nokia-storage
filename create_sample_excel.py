"""
Creates a sample Excel file to test importing phones.
Run: python create_sample_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Nokia Phones"

headers = ["ID", "Name", "Release Date", "Appearance Condition",
           "Working Condition", "Remarks"]
header_fill = PatternFill(start_color="0050C8", end_color="0050C8", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Sample data
phones = [
    ("N3310-001", "Nokia 3310", "2000", "Excellent", "Working", "Classic brick phone"),
    ("N6600-001", "Nokia 6600", "2003", "Good", "Working", "Symbian smartphone"),
    ("N1100-001", "Nokia 1100", "2003", "Good", "Working", "Best seller of all time"),
    ("N8210-001", "Nokia 8210", "1999", "Fair", "Working", "Compact fashion phone"),
    ("N5110-001", "Nokia 5110", "1998", "Good", "Working", "Changeable covers"),
    ("N6110-001", "Nokia 6110", "1997", "Fair", "Partial", "First Nokia with Snake"),
    ("N3210-001", "Nokia 3210", "1999", "Excellent", "Working", "Internal antenna"),
    ("N7110-001", "Nokia 7110", "1999", "Good", "Working", "First WAP phone"),
    ("N8850-001", "Nokia 8850", "1999", "Excellent", "Working", "Premium sliding design"),
    ("N3310-002", "Nokia 3310", "2000", "Fair", "Not Working", "Second unit - for parts"),
    ("N6600-002", "Nokia 6600", "2003", "Poor", "Not Working", "Screen cracked"),
    ("NE71-001", "Nokia E71", "2008", "Good", "Working", "QWERTY business phone"),
    ("NN95-001", "Nokia N95", "2007", "Good", "Working", "Dual slider multimedia"),
    ("N1200-001", "Nokia 1200", "2007", "Excellent", "Working", "Budget flashlight phone"),
    ("N8800-001", "Nokia 8800", "2005", "Excellent", "Working", "Stainless steel body"),
]

for i, (pid, name, date, appear, working, remarks) in enumerate(phones, 2):
    ws.cell(row=i, column=1, value=pid)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=date)
    ws.cell(row=i, column=4, value=appear)
    ws.cell(row=i, column=5, value=working)
    ws.cell(row=i, column=6, value=remarks)

# Auto-width
for col in range(1, 7):
    ws.column_dimensions[chr(64 + col)].width = 22

wb.save("sample_nokia_phones.xlsx")
print("Created: sample_nokia_phones.xlsx")
