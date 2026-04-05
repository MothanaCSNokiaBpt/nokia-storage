"""
Nokia Storage Manager - Complete Android Application
Manage Nokia phones inventory and spare parts with images,
Excel import/export, search, and backup/restore.
"""

import os
import json
import shutil
import zipfile
from datetime import datetime
from functools import partial

from kivy.app import App
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.graphics import Color, Rectangle, RoundedRectangle, Ellipse
from kivy.lang import Builder
from kivy.metrics import dp, sp
from kivy.properties import (
    StringProperty, ListProperty, ObjectProperty,
    NumericProperty, BooleanProperty
)
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image, AsyncImage
from kivy.uix.label import Label
from kivy.uix.modalview import ModalView
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivy.utils import platform

from database import NokiaDatabase

# ── Platform-specific imports ───────────────────────────────────
if platform == "android":
    from android.permissions import request_permissions, Permission, check_permission
    from android.storage import primary_external_storage_path
    from android import mActivity
    from jnius import autoclass, cast
    Environment = autoclass("android.os.Environment")
    Intent = autoclass("android.content.Intent")
    Uri = autoclass("android.net.Uri")
    FileProvider = autoclass("androidx.core.content.FileProvider")
    PythonActivity = autoclass("org.kivy.android.PythonActivity")
    MediaStore = autoclass("android.provider.MediaStore")

# ── Constants ───────────────────────────────────────────────────
NOKIA_BLUE = "#0050C8"
NOKIA_DARK = "#001F6B"
NOKIA_LIGHT = "#F0F5FF"
NOKIA_ACCENT = "#00B5FF"
WHITE = "#FFFFFF"
GREY = "#E0E0E0"
DARK_TEXT = "#1A1A2E"
LIGHT_TEXT = "#666666"
DANGER = "#E53935"
SUCCESS = "#43A047"
WARNING = "#FF9800"

# ── KV Layout ──────────────────────────────────────────────────
KV = """
#:import dp kivy.metrics.dp
#:import sp kivy.metrics.sp
#:import Window kivy.core.window.Window

<RoundedButton@ButtonBehavior+BoxLayout>:
    size_hint_y: None
    height: dp(44)
    padding: dp(16), dp(8)
    canvas.before:
        Color:
            rgba: self.bg_color if hasattr(self, 'bg_color') else (0, 0.314, 0.784, 1)
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(8)]
    bg_color: 0, 0.314, 0.784, 1
    Label:
        text: root.btn_text if hasattr(root, 'btn_text') else ''
        color: 1, 1, 1, 1
        font_size: sp(14)
        bold: True

<IconBtn@ButtonBehavior+BoxLayout>:
    size_hint: None, None
    size: dp(44), dp(44)
    padding: dp(8)
    Image:
        source: root.icon_src if hasattr(root, 'icon_src') else ''
        size: dp(28), dp(28)
        size_hint: None, None
        pos_hint: {'center_x': .5, 'center_y': .5}

<PhoneCard>:
    size_hint_y: None
    height: dp(90)
    padding: dp(8)
    spacing: dp(10)
    orientation: 'horizontal'
    canvas.before:
        Color:
            rgba: 1, 1, 1, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(12)]
        Color:
            rgba: 0.88, 0.88, 0.88, 1
        Line:
            rounded_rectangle: self.x, self.y, self.width, self.height, dp(12)
            width: 1

    AsyncImage:
        source: root.phone_image
        size_hint: None, None
        size: dp(70), dp(70)
        pos_hint: {'center_y': .5}
        allow_stretch: True
        keep_ratio: True

    BoxLayout:
        orientation: 'vertical'
        spacing: dp(2)
        padding: 0, dp(4)
        Label:
            text: root.phone_name
            font_size: sp(16)
            bold: True
            color: 0.1, 0.1, 0.18, 1
            text_size: self.size
            halign: 'left'
            valign: 'middle'
            size_hint_y: 0.4
        Label:
            text: 'ID: ' + root.phone_id
            font_size: sp(12)
            color: 0.4, 0.4, 0.4, 1
            text_size: self.size
            halign: 'left'
            valign: 'middle'
            size_hint_y: 0.3
        Label:
            text: root.phone_date
            font_size: sp(11)
            color: 0.5, 0.5, 0.5, 1
            text_size: self.size
            halign: 'left'
            valign: 'middle'
            size_hint_y: 0.3

<SpareCard>:
    size_hint_y: None
    height: dp(80)
    padding: dp(8)
    spacing: dp(10)
    orientation: 'horizontal'
    canvas.before:
        Color:
            rgba: 1, 1, 1, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(12)]
        Color:
            rgba: 0.88, 0.88, 0.88, 1
        Line:
            rounded_rectangle: self.x, self.y, self.width, self.height, dp(12)
            width: 1

    AsyncImage:
        source: root.spare_image
        size_hint: None, None
        size: dp(60), dp(60)
        pos_hint: {'center_y': .5}
        allow_stretch: True
        keep_ratio: True

    BoxLayout:
        orientation: 'vertical'
        spacing: dp(2)
        padding: 0, dp(4)
        Label:
            text: root.spare_name
            font_size: sp(15)
            bold: True
            color: 0.1, 0.1, 0.18, 1
            text_size: self.size
            halign: 'left'
            valign: 'middle'
            size_hint_y: 0.5
        Label:
            text: root.spare_desc
            font_size: sp(12)
            color: 0.5, 0.5, 0.5, 1
            text_size: self.size
            halign: 'left'
            valign: 'middle'
            size_hint_y: 0.5

<SearchBar>:
    size_hint_y: None
    height: dp(50)
    padding: dp(12), dp(6)
    spacing: dp(8)
    canvas.before:
        Color:
            rgba: 0.95, 0.96, 0.98, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(0)]

    BoxLayout:
        canvas.before:
            Color:
                rgba: 1, 1, 1, 1
            RoundedRectangle:
                pos: self.pos
                size: self.size
                radius: [dp(22)]
            Color:
                rgba: 0.85, 0.85, 0.85, 1
            Line:
                rounded_rectangle: self.x, self.y, self.width, self.height, dp(22)
                width: 1
        padding: dp(14), dp(4)
        spacing: dp(8)
        Label:
            text: chr(0x1F50D)
            size_hint_x: None
            width: dp(30)
            font_size: sp(18)
            color: 0.5, 0.5, 0.5, 1
        TextInput:
            id: search_input
            hint_text: 'Search by name, ID, or date...'
            multiline: False
            background_color: 0, 0, 0, 0
            foreground_color: 0.1, 0.1, 0.1, 1
            hint_text_color: 0.6, 0.6, 0.6, 1
            cursor_color: 0, 0.314, 0.784, 1
            font_size: sp(14)
            padding: 0, dp(8)
            on_text: root.on_search(self.text)

ScreenManager:
    id: sm
    MainScreen:
        name: 'main'
    PhoneDetailScreen:
        name: 'phone_detail'
    AddPhoneScreen:
        name: 'add_phone'
    AddSpareScreen:
        name: 'add_spare'
    ImportExcelScreen:
        name: 'import_excel'
    BulkImageScreen:
        name: 'bulk_images'
    BackupScreen:
        name: 'backup'
    SearchAllScreen:
        name: 'search_all'

<MainScreen>:
    BoxLayout:
        orientation: 'vertical'

        # ── Header ──
        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(16), dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Label:
                text: 'N O K I A  Storage'
                font_size: sp(20)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'
            ButtonBehavior+Label:
                text: chr(0x2630)
                size_hint_x: None
                width: dp(40)
                font_size: sp(24)
                color: 1, 1, 1, 1
                on_release: root.show_menu()

        # ── Search Bar ──
        SearchBar:
            id: search_bar

        # ── Tab Buttons ──
        BoxLayout:
            size_hint_y: None
            height: dp(44)
            canvas.before:
                Color:
                    rgba: 1, 1, 1, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+BoxLayout:
                id: tab_phones
                padding: dp(8)
                on_release: root.switch_tab('phones')
                canvas.before:
                    Color:
                        rgba: (0, 0.314, 0.784, 1) if root.current_tab == 'phones' else (0.95, 0.95, 0.95, 1)
                    Rectangle:
                        pos: self.pos
                        size: self.size
                Label:
                    text: 'Nokia Phones'
                    bold: True
                    font_size: sp(14)
                    color: (1,1,1,1) if root.current_tab == 'phones' else (0.3,0.3,0.3,1)
            ButtonBehavior+BoxLayout:
                id: tab_spares
                padding: dp(8)
                on_release: root.switch_tab('spares')
                canvas.before:
                    Color:
                        rgba: (0, 0.314, 0.784, 1) if root.current_tab == 'spares' else (0.95, 0.95, 0.95, 1)
                    Rectangle:
                        pos: self.pos
                        size: self.size
                Label:
                    text: 'Accessories & Spare Parts'
                    bold: True
                    font_size: sp(14)
                    color: (1,1,1,1) if root.current_tab == 'spares' else (0.3,0.3,0.3,1)

        # ── Content Area ──
        ScrollView:
            id: scroll_view
            do_scroll_x: False
            GridLayout:
                id: content_list
                cols: 1
                spacing: dp(8)
                padding: dp(12)
                size_hint_y: None
                height: self.minimum_height

        # ── Bottom Bar with Add button ──
        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(12), dp(6)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 1, 1, 1, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
                Color:
                    rgba: 0.9, 0.9, 0.9, 1
                Line:
                    points: self.x, self.top, self.right, self.top
                    width: 1
            Label:
                id: count_label
                text: '0 items'
                font_size: sp(12)
                color: 0.5, 0.5, 0.5, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'
            ButtonBehavior+BoxLayout:
                size_hint_x: None
                width: dp(110)
                padding: dp(12), dp(6)
                canvas.before:
                    Color:
                        rgba: 0, 0.314, 0.784, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(22)]
                on_release: root.add_item()
                Label:
                    text: '+ Add'
                    color: 1, 1, 1, 1
                    font_size: sp(14)
                    bold: True
            ButtonBehavior+BoxLayout:
                size_hint_x: None
                width: dp(110)
                padding: dp(12), dp(6)
                canvas.before:
                    Color:
                        rgba: 0, 0.44, 1, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(22)]
                on_release: root.search_all()
                Label:
                    text: 'Search All'
                    color: 1, 1, 1, 1
                    font_size: sp(13)
                    bold: True

<PhoneDetailScreen>:
    BoxLayout:
        orientation: 'vertical'

        # Header
        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Phone Details'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(50)
                text: 'Edit'
                font_size: sp(14)
                color: 1, 1, 1, 1
                on_release: root.edit_phone()
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(50)
                text: 'Del'
                font_size: sp(14)
                color: 1, 0.6, 0.6, 1
                on_release: root.delete_phone()

        ScrollView:
            do_scroll_x: False
            BoxLayout:
                orientation: 'vertical'
                size_hint_y: None
                height: self.minimum_height
                padding: dp(16)
                spacing: dp(12)

                # Phone Image
                BoxLayout:
                    size_hint_y: None
                    height: dp(250)
                    padding: dp(20)
                    canvas.before:
                        Color:
                            rgba: 0.94, 0.96, 1, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(16)]
                    AsyncImage:
                        id: phone_image
                        source: root.image_source
                        allow_stretch: True
                        keep_ratio: True

                # Add/Change Image Button
                ButtonBehavior+BoxLayout:
                    size_hint_y: None
                    height: dp(38)
                    padding: dp(12), dp(6)
                    canvas.before:
                        Color:
                            rgba: 0, 0.314, 0.784, 0.1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(8)]
                    on_release: root.change_image()
                    Label:
                        text: 'Change Image'
                        color: 0, 0.314, 0.784, 1
                        font_size: sp(13)
                        bold: True

                # Info Card
                BoxLayout:
                    orientation: 'vertical'
                    size_hint_y: None
                    height: self.minimum_height
                    padding: dp(16)
                    spacing: dp(10)
                    canvas.before:
                        Color:
                            rgba: 1, 1, 1, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(12)]

                    Label:
                        id: lbl_name
                        text: root.p_name
                        font_size: sp(22)
                        bold: True
                        color: 0.1, 0.1, 0.18, 1
                        size_hint_y: None
                        height: dp(32)
                        text_size: self.size
                        halign: 'left'
                    Label:
                        id: lbl_id
                        text: 'ID: ' + root.p_id
                        font_size: sp(14)
                        color: 0.4, 0.4, 0.4, 1
                        size_hint_y: None
                        height: dp(22)
                        text_size: self.size
                        halign: 'left'
                    Label:
                        text: 'Release: ' + root.p_date
                        font_size: sp(14)
                        color: 0.4, 0.4, 0.4, 1
                        size_hint_y: None
                        height: dp(22)
                        text_size: self.size
                        halign: 'left'
                    # Conditions
                    BoxLayout:
                        size_hint_y: None
                        height: dp(30)
                        spacing: dp(8)
                        BoxLayout:
                            padding: dp(8), dp(4)
                            canvas.before:
                                Color:
                                    rgba: 0.26, 0.63, 0.28, 0.15
                                RoundedRectangle:
                                    pos: self.pos
                                    size: self.size
                                    radius: [dp(6)]
                            Label:
                                text: 'Look: ' + root.p_appear
                                font_size: sp(12)
                                color: 0.26, 0.63, 0.28, 1
                                bold: True
                        BoxLayout:
                            padding: dp(8), dp(4)
                            canvas.before:
                                Color:
                                    rgba: 0, 0.314, 0.784, 0.15
                                RoundedRectangle:
                                    pos: self.pos
                                    size: self.size
                                    radius: [dp(6)]
                            Label:
                                text: 'Work: ' + root.p_working
                                font_size: sp(12)
                                color: 0, 0.314, 0.784, 1
                                bold: True
                    Label:
                        text: 'Remarks:'
                        font_size: sp(13)
                        bold: True
                        color: 0.3, 0.3, 0.3, 1
                        size_hint_y: None
                        height: dp(20)
                        text_size: self.size
                        halign: 'left'
                    Label:
                        text: root.p_remarks
                        font_size: sp(13)
                        color: 0.4, 0.4, 0.4, 1
                        size_hint_y: None
                        height: self.texture_size[1] + dp(10)
                        text_size: self.width, None
                        halign: 'left'

                # Spare Parts Section
                Label:
                    text: 'Related Spare Parts'
                    font_size: sp(16)
                    bold: True
                    color: 0.1, 0.1, 0.18, 1
                    size_hint_y: None
                    height: dp(30)
                    text_size: self.size
                    halign: 'left'

                GridLayout:
                    id: spare_parts_grid
                    cols: 1
                    spacing: dp(8)
                    size_hint_y: None
                    height: self.minimum_height

                # Add Spare Part Button
                ButtonBehavior+BoxLayout:
                    size_hint_y: None
                    height: dp(40)
                    padding: dp(12), dp(8)
                    canvas.before:
                        Color:
                            rgba: 0, 0.314, 0.784, 0.1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(8)]
                    on_release: root.add_spare_for_phone()
                    Label:
                        text: '+ Add Spare Part'
                        color: 0, 0.314, 0.784, 1
                        font_size: sp(14)
                        bold: True

                Widget:
                    size_hint_y: None
                    height: dp(40)

<AddPhoneScreen>:
    BoxLayout:
        orientation: 'vertical'

        # Header
        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: root.screen_title
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        ScrollView:
            do_scroll_x: False
            BoxLayout:
                orientation: 'vertical'
                size_hint_y: None
                height: self.minimum_height
                padding: dp(16)
                spacing: dp(14)

                # Image Preview
                BoxLayout:
                    size_hint_y: None
                    height: dp(180)
                    padding: dp(16)
                    canvas.before:
                        Color:
                            rgba: 0.94, 0.96, 1, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(12)]
                    AsyncImage:
                        id: preview_image
                        source: root.image_preview
                        allow_stretch: True
                        keep_ratio: True

                BoxLayout:
                    size_hint_y: None
                    height: dp(40)
                    spacing: dp(8)
                    ButtonBehavior+BoxLayout:
                        padding: dp(8), dp(6)
                        canvas.before:
                            Color:
                                rgba: 0, 0.314, 0.784, 0.15
                            RoundedRectangle:
                                pos: self.pos
                                size: self.size
                                radius: [dp(8)]
                        on_release: root.pick_from_gallery()
                        Label:
                            text: 'Gallery'
                            color: 0, 0.314, 0.784, 1
                            font_size: sp(13)
                            bold: True
                    ButtonBehavior+BoxLayout:
                        padding: dp(8), dp(6)
                        canvas.before:
                            Color:
                                rgba: 0, 0.314, 0.784, 0.15
                            RoundedRectangle:
                                pos: self.pos
                                size: self.size
                                radius: [dp(8)]
                        on_release: root.take_photo()
                        Label:
                            text: 'Camera'
                            color: 0, 0.314, 0.784, 1
                            font_size: sp(13)
                            bold: True

                # Form Fields
                Label:
                    text: 'Phone ID *'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_id
                    hint_text: 'e.g. NOKIA-3310-001'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1
                    cursor_color: 0, 0.314, 0.784, 1

                Label:
                    text: 'Phone Name *'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_name
                    hint_text: 'e.g. Nokia 3310'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Release Date'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_date
                    hint_text: 'e.g. 2000'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Appearance Condition'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_appear
                    hint_text: 'e.g. Excellent / Good / Fair / Poor'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Working Condition'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_working
                    hint_text: 'e.g. Working / Not Working / Partial'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Remarks'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: input_remarks
                    hint_text: 'Any notes...'
                    multiline: True
                    size_hint_y: None
                    height: dp(80)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                # Save Button
                ButtonBehavior+BoxLayout:
                    size_hint_y: None
                    height: dp(48)
                    padding: dp(16), dp(10)
                    canvas.before:
                        Color:
                            rgba: 0, 0.314, 0.784, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(10)]
                    on_release: root.save_phone()
                    Label:
                        text: 'Save Phone'
                        color: 1, 1, 1, 1
                        font_size: sp(16)
                        bold: True

                Widget:
                    size_hint_y: None
                    height: dp(30)

<AddSpareScreen>:
    BoxLayout:
        orientation: 'vertical'

        # Header
        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Add Spare Part'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        ScrollView:
            do_scroll_x: False
            BoxLayout:
                orientation: 'vertical'
                size_hint_y: None
                height: self.minimum_height
                padding: dp(16)
                spacing: dp(14)

                # Image Preview
                BoxLayout:
                    size_hint_y: None
                    height: dp(180)
                    padding: dp(16)
                    canvas.before:
                        Color:
                            rgba: 0.94, 0.96, 1, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(12)]
                    AsyncImage:
                        id: spare_preview_image
                        source: root.image_preview
                        allow_stretch: True
                        keep_ratio: True

                BoxLayout:
                    size_hint_y: None
                    height: dp(40)
                    spacing: dp(8)
                    ButtonBehavior+BoxLayout:
                        padding: dp(8), dp(6)
                        canvas.before:
                            Color:
                                rgba: 0, 0.314, 0.784, 0.15
                            RoundedRectangle:
                                pos: self.pos
                                size: self.size
                                radius: [dp(8)]
                        on_release: root.pick_from_gallery()
                        Label:
                            text: 'Gallery'
                            color: 0, 0.314, 0.784, 1
                            font_size: sp(13)
                            bold: True
                    ButtonBehavior+BoxLayout:
                        padding: dp(8), dp(6)
                        canvas.before:
                            Color:
                                rgba: 0, 0.314, 0.784, 0.15
                            RoundedRectangle:
                                pos: self.pos
                                size: self.size
                                radius: [dp(8)]
                        on_release: root.take_photo()
                        Label:
                            text: 'Camera'
                            color: 0, 0.314, 0.784, 1
                            font_size: sp(13)
                            bold: True

                Label:
                    text: 'Spare Part Name *'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: spare_input_name
                    hint_text: 'Same as phone name (e.g. Nokia 3310)'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Description'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: spare_input_desc
                    hint_text: 'e.g. Battery cover, Screen, etc.'
                    multiline: True
                    size_hint_y: None
                    height: dp(80)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                Label:
                    text: 'Link to Phone ID (optional)'
                    font_size: sp(13)
                    color: 0.3, 0.3, 0.3, 1
                    size_hint_y: None
                    height: dp(20)
                    text_size: self.size
                    halign: 'left'
                TextInput:
                    id: spare_input_phone_id
                    hint_text: 'e.g. NOKIA-3310-001'
                    multiline: False
                    size_hint_y: None
                    height: dp(44)
                    font_size: sp(14)
                    padding: dp(12), dp(10)
                    background_normal: ''
                    background_active: ''
                    background_color: 0.97, 0.97, 0.97, 1
                    foreground_color: 0.1, 0.1, 0.1, 1

                # Save Button
                ButtonBehavior+BoxLayout:
                    size_hint_y: None
                    height: dp(48)
                    padding: dp(16), dp(10)
                    canvas.before:
                        Color:
                            rgba: 0, 0.314, 0.784, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(10)]
                    on_release: root.save_spare()
                    Label:
                        text: 'Save Spare Part'
                        color: 1, 1, 1, 1
                        font_size: sp(16)
                        bold: True

                Widget:
                    size_hint_y: None
                    height: dp(30)

<ImportExcelScreen>:
    BoxLayout:
        orientation: 'vertical'

        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Import from Excel'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        BoxLayout:
            orientation: 'vertical'
            padding: dp(20)
            spacing: dp(16)

            Label:
                text: 'Select an Excel file (.xlsx) containing phone data.\\n\\nExpected columns:\\nID | Name | Release Date | Appearance | Working | Remarks'
                font_size: sp(14)
                color: 0.3, 0.3, 0.3, 1
                text_size: self.width - dp(20), None
                size_hint_y: None
                height: self.texture_size[1] + dp(20)
                halign: 'left'

            Label:
                id: import_status
                text: ''
                font_size: sp(14)
                color: 0.26, 0.63, 0.28, 1
                size_hint_y: None
                height: dp(30)

            ButtonBehavior+BoxLayout:
                size_hint_y: None
                height: dp(50)
                padding: dp(16), dp(12)
                canvas.before:
                    Color:
                        rgba: 0, 0.314, 0.784, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(10)]
                on_release: root.select_file()
                Label:
                    text: 'Select Excel File'
                    color: 1, 1, 1, 1
                    font_size: sp(16)
                    bold: True

            Widget:

<BulkImageScreen>:
    BoxLayout:
        orientation: 'vertical'

        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Bulk Image Import'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        BoxLayout:
            orientation: 'vertical'
            padding: dp(16)
            spacing: dp(12)

            # Target selection
            BoxLayout:
                size_hint_y: None
                height: dp(40)
                spacing: dp(8)
                ButtonBehavior+BoxLayout:
                    padding: dp(12), dp(6)
                    canvas.before:
                        Color:
                            rgba: (0, 0.314, 0.784, 1) if root.target_type == 'phones' else (0.9, 0.9, 0.9, 1)
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(8)]
                    on_release: root.set_target('phones')
                    Label:
                        text: 'For Phones'
                        color: (1,1,1,1) if root.target_type == 'phones' else (0.3,0.3,0.3,1)
                        font_size: sp(13)
                        bold: True
                ButtonBehavior+BoxLayout:
                    padding: dp(12), dp(6)
                    canvas.before:
                        Color:
                            rgba: (0, 0.314, 0.784, 1) if root.target_type == 'spares' else (0.9, 0.9, 0.9, 1)
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [dp(8)]
                    on_release: root.set_target('spares')
                    Label:
                        text: 'For Spare Parts'
                        color: (1,1,1,1) if root.target_type == 'spares' else (0.3,0.3,0.3,1)
                        font_size: sp(13)
                        bold: True

            ButtonBehavior+BoxLayout:
                size_hint_y: None
                height: dp(44)
                padding: dp(12), dp(8)
                canvas.before:
                    Color:
                        rgba: 0, 0.314, 0.784, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(10)]
                on_release: root.select_images()
                Label:
                    text: 'Select Images from Gallery'
                    color: 1, 1, 1, 1
                    font_size: sp(14)
                    bold: True

            Label:
                id: bulk_status
                text: 'Select images to assign to phones or spare parts'
                font_size: sp(13)
                color: 0.5, 0.5, 0.5, 1
                size_hint_y: None
                height: dp(24)

            ScrollView:
                do_scroll_x: False
                GridLayout:
                    id: bulk_grid
                    cols: 1
                    spacing: dp(10)
                    size_hint_y: None
                    height: self.minimum_height

<BackupScreen>:
    BoxLayout:
        orientation: 'vertical'

        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Backup & Restore'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        BoxLayout:
            orientation: 'vertical'
            padding: dp(20)
            spacing: dp(16)

            Label:
                text: 'Create a full backup of all data and images.\\nRestore on a new device after installing the app.'
                font_size: sp(14)
                color: 0.3, 0.3, 0.3, 1
                text_size: self.width - dp(20), None
                size_hint_y: None
                height: self.texture_size[1] + dp(10)
                halign: 'left'

            ButtonBehavior+BoxLayout:
                size_hint_y: None
                height: dp(50)
                padding: dp(16), dp(12)
                canvas.before:
                    Color:
                        rgba: 0.26, 0.63, 0.28, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(10)]
                on_release: root.create_backup()
                Label:
                    text: 'Create Backup'
                    color: 1, 1, 1, 1
                    font_size: sp(16)
                    bold: True

            ButtonBehavior+BoxLayout:
                size_hint_y: None
                height: dp(50)
                padding: dp(16), dp(12)
                canvas.before:
                    Color:
                        rgba: 1, 0.6, 0, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(10)]
                on_release: root.restore_backup()
                Label:
                    text: 'Restore from Backup'
                    color: 1, 1, 1, 1
                    font_size: sp(16)
                    bold: True

            ButtonBehavior+BoxLayout:
                size_hint_y: None
                height: dp(50)
                padding: dp(16), dp(12)
                canvas.before:
                    Color:
                        rgba: 0, 0.314, 0.784, 1
                    RoundedRectangle:
                        pos: self.pos
                        size: self.size
                        radius: [dp(10)]
                on_release: root.share_backup()
                Label:
                    text: 'Share Backup (Email/Drive)'
                    color: 1, 1, 1, 1
                    font_size: sp(16)
                    bold: True

            Label:
                id: backup_status
                text: ''
                font_size: sp(14)
                color: 0.26, 0.63, 0.28, 1
                size_hint_y: None
                height: dp(30)

            Widget:

<SearchAllScreen>:
    BoxLayout:
        orientation: 'vertical'

        BoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            canvas.before:
                Color:
                    rgba: 0, 0.314, 0.784, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            ButtonBehavior+Label:
                size_hint_x: None
                width: dp(40)
                text: '<'
                font_size: sp(24)
                bold: True
                color: 1, 1, 1, 1
                on_release: root.go_back()
            Label:
                text: 'Search All'
                font_size: sp(18)
                bold: True
                color: 1, 1, 1, 1
                text_size: self.size
                halign: 'left'
                valign: 'middle'

        SearchBar:
            id: search_all_bar

        ScrollView:
            do_scroll_x: False
            GridLayout:
                id: results_list
                cols: 1
                spacing: dp(8)
                padding: dp(12)
                size_hint_y: None
                height: self.minimum_height
"""


# ── Helper to get app storage paths ────────────────────────────
def get_app_path():
    if platform == "android":
        from android.storage import app_storage_path
        return app_storage_path()
    return os.path.dirname(os.path.abspath(__file__))


def get_images_path():
    p = os.path.join(get_app_path(), "images")
    os.makedirs(p, exist_ok=True)
    return p


def get_phone_images_path():
    p = os.path.join(get_images_path(), "phones")
    os.makedirs(p, exist_ok=True)
    return p


def get_spare_images_path():
    p = os.path.join(get_images_path(), "spares")
    os.makedirs(p, exist_ok=True)
    return p


def get_db_path():
    return os.path.join(get_app_path(), "nokia_storage.db")


def get_backup_path():
    if platform == "android":
        return os.path.join(
            primary_external_storage_path(), "Download"
        )
    return os.path.join(get_app_path(), "backups")


def get_default_image():
    """Return path to default phone image, create if needed."""
    p = os.path.join(get_app_path(), "default_phone.png")
    if not os.path.exists(p):
        try:
            from PIL import Image as PILImage, ImageDraw, ImageFont
            img = PILImage.new("RGB", (200, 200), (240, 245, 255))
            draw = ImageDraw.Draw(img)
            draw.rounded_rectangle([30, 20, 170, 180], radius=15,
                                    fill=(200, 210, 230), outline=(150, 160, 180))
            draw.rounded_rectangle([55, 50, 145, 120], radius=5,
                                    fill=(180, 195, 220))
            draw.ellipse([85, 135, 115, 155], fill=(180, 195, 220))
            img.save(p)
        except Exception:
            return ""
    return p


def copy_image_to_storage(source_path, dest_folder):
    """Copy an image to app storage and return new path."""
    if not source_path or not os.path.exists(source_path):
        return ""
    ext = os.path.splitext(source_path)[1] or ".jpg"
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{ext}"
    dest = os.path.join(dest_folder, filename)
    try:
        shutil.copy2(source_path, dest)
        return dest
    except Exception:
        return ""


# ── Custom Widgets ──────────────────────────────────────────────

class PhoneCard(ButtonBehavior, BoxLayout):
    phone_id = StringProperty("")
    phone_name = StringProperty("")
    phone_date = StringProperty("")
    phone_image = StringProperty("")


class SpareCard(ButtonBehavior, BoxLayout):
    spare_id = NumericProperty(0)
    spare_name = StringProperty("")
    spare_desc = StringProperty("")
    spare_image = StringProperty("")


class SearchBar(BoxLayout):
    def on_search(self, text):
        app = App.get_running_app()
        screen = app.root.current_screen
        if hasattr(screen, "do_search"):
            screen.do_search(text)


# ── Screens ─────────────────────────────────────────────────────

class MainScreen(Screen):
    current_tab = StringProperty("phones")

    def on_enter(self):
        self.refresh_list()

    def switch_tab(self, tab):
        self.current_tab = tab
        self.ids.search_bar.ids.search_input.text = ""
        self.refresh_list()

    def refresh_list(self):
        app = App.get_running_app()
        grid = self.ids.content_list
        grid.clear_widgets()
        default_img = get_default_image()

        if self.current_tab == "phones":
            phones = app.db.get_all_phones()
            self.ids.count_label.text = f"{len(phones)} phones"
            for p in phones:
                img = p.get("image_path", "") or default_img
                card = PhoneCard(
                    phone_id=p["id"],
                    phone_name=p["name"],
                    phone_date=p.get("release_date", ""),
                    phone_image=img,
                )
                card.bind(on_release=partial(self._open_phone, p["id"]))
                grid.add_widget(card)
        else:
            spares = app.db.get_all_spare_parts()
            self.ids.count_label.text = f"{len(spares)} spare parts"
            for s in spares:
                img = s.get("image_path", "") or default_img
                card = SpareCard(
                    spare_id=s["id"],
                    spare_name=s["name"],
                    spare_desc=s.get("description", ""),
                    spare_image=img,
                )
                card.bind(on_release=partial(self._open_spare_detail, s["id"]))
                grid.add_widget(card)

    def do_search(self, text):
        app = App.get_running_app()
        grid = self.ids.content_list
        grid.clear_widgets()
        default_img = get_default_image()

        if not text.strip():
            self.refresh_list()
            return

        if self.current_tab == "phones":
            results = app.db.search_phones(text)
            self.ids.count_label.text = f"{len(results)} found"
            for p in results:
                img = p.get("image_path", "") or default_img
                card = PhoneCard(
                    phone_id=p["id"],
                    phone_name=p["name"],
                    phone_date=p.get("release_date", ""),
                    phone_image=img,
                )
                card.bind(on_release=partial(self._open_phone, p["id"]))
                grid.add_widget(card)
        else:
            results = app.db.search_spare_parts(text)
            self.ids.count_label.text = f"{len(results)} found"
            for s in results:
                img = s.get("image_path", "") or default_img
                card = SpareCard(
                    spare_id=s["id"],
                    spare_name=s["name"],
                    spare_desc=s.get("description", ""),
                    spare_image=img,
                )
                card.bind(on_release=partial(self._open_spare_detail, s["id"]))
                grid.add_widget(card)

    def _open_phone(self, phone_id, *args):
        app = App.get_running_app()
        detail = app.root.get_screen("phone_detail")
        detail.load_phone(phone_id)
        app.root.transition = SlideTransition(direction="left")
        app.root.current = "phone_detail"

    def _open_spare_detail(self, spare_id, *args):
        # For spare parts, show a simple popup with image and details
        app = App.get_running_app()
        spare = app.db.get_spare_part(spare_id)
        if not spare:
            return
        self._show_spare_popup(spare)

    def _show_spare_popup(self, spare):
        default_img = get_default_image()
        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(16))
        img = AsyncImage(
            source=spare.get("image_path", "") or default_img,
            size_hint_y=None, height=dp(200),
            allow_stretch=True, keep_ratio=True,
        )
        content.add_widget(img)
        content.add_widget(Label(
            text=spare["name"], font_size=sp(18), bold=True,
            color=(0.1, 0.1, 0.18, 1), size_hint_y=None, height=dp(30),
        ))
        if spare.get("description"):
            content.add_widget(Label(
                text=spare["description"], font_size=sp(14),
                color=(0.4, 0.4, 0.4, 1), size_hint_y=None, height=dp(24),
            ))

        # Delete button
        del_btn = ButtonBehavior.__class__.__mro__[0]  # workaround
        popup = ModalView(size_hint=(0.85, None), height=dp(360))
        popup.add_widget(content)
        popup.open()

    def add_item(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="left")
        if self.current_tab == "phones":
            screen = app.root.get_screen("add_phone")
            screen.edit_mode = False
            screen.clear_form()
            app.root.current = "add_phone"
        else:
            screen = app.root.get_screen("add_spare")
            screen.clear_form()
            app.root.current = "add_spare"

    def search_all(self):
        app = App.get_running_app()
        query = self.ids.search_bar.ids.search_input.text
        search_screen = app.root.get_screen("search_all")
        search_screen.initial_query = query
        app.root.transition = SlideTransition(direction="left")
        app.root.current = "search_all"

    def show_menu(self):
        app = App.get_running_app()
        content = BoxLayout(
            orientation="vertical", spacing=dp(4), padding=dp(8),
            size_hint_y=None,
        )
        items = [
            ("Import from Excel", lambda *a: self._goto("import_excel")),
            ("Export to Excel", lambda *a: app.export_to_excel()),
            ("Bulk Image Import", lambda *a: self._goto("bulk_images")),
            ("Backup & Restore", lambda *a: self._goto("backup")),
        ]
        content.height = dp(48) * len(items)

        popup = ModalView(
            size_hint=(0.7, None), height=dp(48) * len(items) + dp(20),
            pos_hint={"top": 0.92, "right": 0.95},
        )

        for label_text, callback in items:
            btn = ButtonBehavior.__new__(type("Btn", (ButtonBehavior, BoxLayout), {}))
            box = BoxLayout(
                size_hint_y=None, height=dp(44), padding=(dp(12), dp(6)),
            )
            lbl = Label(
                text=label_text, font_size=sp(14),
                color=(0.1, 0.1, 0.18, 1),
                text_size=(dp(200), None), halign="left",
            )
            box.add_widget(lbl)
            box.bind(on_touch_down=lambda w, t, cb=callback, p=popup: (
                cb() or p.dismiss()) if w.collide_point(*t.pos) else None
            )
            content.add_widget(box)

        popup.add_widget(content)
        popup.open()

    def _goto(self, screen_name):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="left")
        app.root.current = screen_name


class PhoneDetailScreen(Screen):
    p_id = StringProperty("")
    p_name = StringProperty("")
    p_date = StringProperty("")
    p_appear = StringProperty("")
    p_working = StringProperty("")
    p_remarks = StringProperty("")
    image_source = StringProperty("")

    def load_phone(self, phone_id):
        app = App.get_running_app()
        phone = app.db.get_phone(phone_id)
        if not phone:
            return
        self.p_id = phone["id"]
        self.p_name = phone["name"]
        self.p_date = phone.get("release_date", "") or ""
        self.p_appear = phone.get("appearance_condition", "") or ""
        self.p_working = phone.get("working_condition", "") or ""
        self.p_remarks = phone.get("remarks", "") or ""
        self.image_source = phone.get("image_path", "") or get_default_image()

        # Load spare parts
        self._load_spare_parts(phone["name"])

    def _load_spare_parts(self, phone_name):
        app = App.get_running_app()
        grid = self.ids.spare_parts_grid
        grid.clear_widgets()
        default_img = get_default_image()
        spares = app.db.get_spare_parts_for_phone(phone_name)

        if not spares:
            lbl = Label(
                text="No spare parts found", font_size=sp(13),
                color=(0.5, 0.5, 0.5, 1), size_hint_y=None, height=dp(30),
            )
            grid.add_widget(lbl)
            return

        for s in spares:
            img = s.get("image_path", "") or default_img
            card = SpareCard(
                spare_id=s["id"],
                spare_name=s["name"],
                spare_desc=s.get("description", ""),
                spare_image=img,
            )
            grid.add_widget(card)

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"

    def edit_phone(self):
        app = App.get_running_app()
        screen = app.root.get_screen("add_phone")
        screen.edit_mode = True
        screen.load_for_edit(self.p_id)
        app.root.transition = SlideTransition(direction="left")
        app.root.current = "add_phone"

    def delete_phone(self):
        app = App.get_running_app()
        content = BoxLayout(
            orientation="vertical", spacing=dp(12), padding=dp(16),
        )
        content.add_widget(Label(
            text=f"Delete {self.p_name}?",
            font_size=sp(16), color=(0.1, 0.1, 0.18, 1),
            size_hint_y=None, height=dp(30),
        ))
        btn_row = BoxLayout(spacing=dp(8), size_hint_y=None, height=dp(44))

        popup = ModalView(size_hint=(0.8, None), height=dp(140))

        cancel_box = BoxLayout(padding=(dp(8), dp(6)))
        cancel_box.add_widget(Label(text="Cancel", font_size=sp(14), color=(0.4, 0.4, 0.4, 1)))
        cancel_box.bind(on_touch_down=lambda w, t: popup.dismiss() if w.collide_point(*t.pos) else None)

        delete_box = BoxLayout(padding=(dp(8), dp(6)))
        with delete_box.canvas.before:
            Color(0.9, 0.22, 0.21, 1)
            delete_box._rrect = RoundedRectangle(pos=delete_box.pos, size=delete_box.size, radius=[dp(8)])
        delete_box.bind(pos=lambda w, v: setattr(w._rrect, "pos", v))
        delete_box.bind(size=lambda w, v: setattr(w._rrect, "size", v))
        delete_box.add_widget(Label(text="Delete", font_size=sp(14), color=(1, 1, 1, 1), bold=True))
        delete_box.bind(on_touch_down=lambda w, t: (
            self._confirm_delete(popup)) if w.collide_point(*t.pos) else None
        )

        btn_row.add_widget(cancel_box)
        btn_row.add_widget(delete_box)
        content.add_widget(btn_row)
        popup.add_widget(content)
        popup.open()

    def _confirm_delete(self, popup):
        app = App.get_running_app()
        # Delete image file if exists
        phone = app.db.get_phone(self.p_id)
        if phone and phone.get("image_path") and os.path.exists(phone["image_path"]):
            try:
                os.remove(phone["image_path"])
            except Exception:
                pass
        app.db.delete_phone(self.p_id)
        popup.dismiss()
        self.go_back()

    def change_image(self):
        app = App.get_running_app()
        app.pick_image_for = ("phone", self.p_id)
        app.open_file_chooser()

    def add_spare_for_phone(self):
        app = App.get_running_app()
        screen = app.root.get_screen("add_spare")
        screen.clear_form()
        screen.ids.spare_input_name.text = self.p_name
        screen.ids.spare_input_phone_id.text = self.p_id
        app.root.transition = SlideTransition(direction="left")
        app.root.current = "add_spare"


class AddPhoneScreen(Screen):
    edit_mode = BooleanProperty(False)
    screen_title = StringProperty("Add Phone")
    image_preview = StringProperty("")
    _selected_image = StringProperty("")

    def on_edit_mode(self, *args):
        self.screen_title = "Edit Phone" if self.edit_mode else "Add Phone"

    def clear_form(self):
        self.image_preview = get_default_image()
        self._selected_image = ""
        Clock.schedule_once(self._clear_inputs, 0.1)

    def _clear_inputs(self, *args):
        try:
            self.ids.input_id.text = ""
            self.ids.input_name.text = ""
            self.ids.input_date.text = ""
            self.ids.input_appear.text = ""
            self.ids.input_working.text = ""
            self.ids.input_remarks.text = ""
        except Exception:
            pass

    def load_for_edit(self, phone_id):
        app = App.get_running_app()
        phone = app.db.get_phone(phone_id)
        if not phone:
            return
        self.image_preview = phone.get("image_path", "") or get_default_image()
        self._selected_image = phone.get("image_path", "")
        Clock.schedule_once(partial(self._fill_inputs, phone), 0.1)

    def _fill_inputs(self, phone, *args):
        self.ids.input_id.text = phone["id"]
        self.ids.input_name.text = phone["name"]
        self.ids.input_date.text = phone.get("release_date", "") or ""
        self.ids.input_appear.text = phone.get("appearance_condition", "") or ""
        self.ids.input_working.text = phone.get("working_condition", "") or ""
        self.ids.input_remarks.text = phone.get("remarks", "") or ""

    def pick_from_gallery(self):
        app = App.get_running_app()
        app.pick_image_for = ("add_phone_screen", None)
        app.open_file_chooser()

    def take_photo(self):
        app = App.get_running_app()
        app.pick_image_for = ("add_phone_screen", None)
        app.take_camera_photo()

    def on_image_selected(self, path):
        self._selected_image = path
        self.image_preview = path

    def save_phone(self):
        app = App.get_running_app()
        phone_id = self.ids.input_id.text.strip()
        name = self.ids.input_name.text.strip()

        if not phone_id or not name:
            app.show_toast("ID and Name are required")
            return

        # Copy image to storage if new
        image_path = self._selected_image
        if image_path and not image_path.startswith(get_phone_images_path()):
            image_path = copy_image_to_storage(image_path, get_phone_images_path())

        app.db.add_phone(
            phone_id=phone_id,
            name=name,
            release_date=self.ids.input_date.text.strip(),
            appearance=self.ids.input_appear.text.strip(),
            working=self.ids.input_working.text.strip(),
            remarks=self.ids.input_remarks.text.strip(),
            image_path=image_path,
        )
        app.show_toast("Phone saved!")
        self.go_back()

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


class AddSpareScreen(Screen):
    image_preview = StringProperty("")
    _selected_image = StringProperty("")

    def clear_form(self):
        self.image_preview = get_default_image()
        self._selected_image = ""
        Clock.schedule_once(self._clear_inputs, 0.1)

    def _clear_inputs(self, *args):
        try:
            self.ids.spare_input_name.text = ""
            self.ids.spare_input_desc.text = ""
            self.ids.spare_input_phone_id.text = ""
        except Exception:
            pass

    def pick_from_gallery(self):
        app = App.get_running_app()
        app.pick_image_for = ("add_spare_screen", None)
        app.open_file_chooser()

    def take_photo(self):
        app = App.get_running_app()
        app.pick_image_for = ("add_spare_screen", None)
        app.take_camera_photo()

    def on_image_selected(self, path):
        self._selected_image = path
        self.image_preview = path

    def save_spare(self):
        app = App.get_running_app()
        name = self.ids.spare_input_name.text.strip()

        if not name:
            app.show_toast("Name is required")
            return

        image_path = self._selected_image
        if image_path and not image_path.startswith(get_spare_images_path()):
            image_path = copy_image_to_storage(image_path, get_spare_images_path())

        app.db.add_spare_part(
            name=name,
            phone_id=self.ids.spare_input_phone_id.text.strip(),
            image_path=image_path,
            description=self.ids.spare_input_desc.text.strip(),
        )
        app.show_toast("Spare part saved!")
        self.go_back()

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


class ImportExcelScreen(Screen):
    def select_file(self):
        app = App.get_running_app()
        app.pick_image_for = ("import_excel", None)
        app.open_file_chooser(filters=["*.xlsx", "*.xls"])

    def on_file_selected(self, path):
        app = App.get_running_app()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active

            rows = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip().lower() if h else "" for h in row]
                    continue
                if not any(row):
                    continue
                data = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        key = headers[j]
                        # Map common column names
                        if key in ("id", "phone_id", "phone id"):
                            data["id"] = str(val) if val else ""
                        elif key in ("name", "phone_name", "phone name", "model"):
                            data["name"] = str(val) if val else ""
                        elif key in ("release_date", "release date", "date", "year"):
                            data["release_date"] = str(val) if val else ""
                        elif key in ("appearance_condition", "appearance condition",
                                     "appearance", "look"):
                            data["appearance_condition"] = str(val) if val else ""
                        elif key in ("working_condition", "working condition",
                                     "working", "status"):
                            data["working_condition"] = str(val) if val else ""
                        elif key in ("remarks", "notes", "comment", "comments"):
                            data["remarks"] = str(val) if val else ""
                if data.get("id") or data.get("name"):
                    if not data.get("id"):
                        data["id"] = f"AUTO-{i}"
                    rows.append(data)
            wb.close()

            count = app.db.import_phones_from_rows(rows)
            self.ids.import_status.text = f"Imported {count} phones!"
            self.ids.import_status.color = (0.26, 0.63, 0.28, 1)
        except Exception as e:
            self.ids.import_status.text = f"Error: {str(e)}"
            self.ids.import_status.color = (0.9, 0.22, 0.21, 1)

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


class BulkImageScreen(Screen):
    target_type = StringProperty("phones")

    def set_target(self, t):
        self.target_type = t

    def select_images(self):
        app = App.get_running_app()
        app.pick_image_for = ("bulk_images", self.target_type)
        app.open_file_chooser(multiple=True)

    def on_images_selected(self, paths):
        grid = self.ids.bulk_grid
        grid.clear_widgets()
        self.ids.bulk_status.text = f"{len(paths)} images selected"

        for path in paths:
            row = BoxLayout(
                orientation="horizontal", size_hint_y=None,
                height=dp(110), spacing=dp(8), padding=dp(4),
            )
            with row.canvas.before:
                Color(1, 1, 1, 1)
                row._bg = RoundedRectangle(pos=row.pos, size=row.size, radius=[dp(8)])
            row.bind(pos=lambda w, v: setattr(w._bg, "pos", v))
            row.bind(size=lambda w, v: setattr(w._bg, "size", v))

            img = AsyncImage(
                source=path, size_hint=(None, 1), width=dp(80),
                allow_stretch=True, keep_ratio=True,
            )
            row.add_widget(img)

            form = BoxLayout(orientation="vertical", spacing=dp(4), padding=(0, dp(4)))
            name_input = TextInput(
                hint_text="Name (e.g. Nokia 3310)",
                multiline=False, size_hint_y=None, height=dp(36),
                font_size=sp(13), padding=(dp(8), dp(6)),
                background_normal="", background_active="",
                background_color=(0.97, 0.97, 0.97, 1),
            )
            form.add_widget(name_input)

            if self.target_type == "phones":
                id_input = TextInput(
                    hint_text="Phone ID",
                    multiline=False, size_hint_y=None, height=dp(36),
                    font_size=sp(13), padding=(dp(8), dp(6)),
                    background_normal="", background_active="",
                    background_color=(0.97, 0.97, 0.97, 1),
                )
                form.add_widget(id_input)
            else:
                id_input = None

            row.add_widget(form)
            row._path = path
            row._name_input = name_input
            row._id_input = id_input
            grid.add_widget(row)

        # Save all button
        save_btn = BoxLayout(size_hint_y=None, height=dp(48), padding=(dp(12), dp(8)))
        with save_btn.canvas.before:
            Color(0, 0.314, 0.784, 1)
            save_btn._bg = RoundedRectangle(pos=save_btn.pos, size=save_btn.size, radius=[dp(10)])
        save_btn.bind(pos=lambda w, v: setattr(w._bg, "pos", v))
        save_btn.bind(size=lambda w, v: setattr(w._bg, "size", v))
        save_btn.add_widget(Label(
            text="Save All", color=(1, 1, 1, 1), font_size=sp(16), bold=True,
        ))
        save_btn.bind(on_touch_down=lambda w, t: self._save_all() if w.collide_point(*t.pos) else None)
        grid.add_widget(save_btn)

    def _save_all(self):
        app = App.get_running_app()
        grid = self.ids.bulk_grid
        count = 0

        for child in grid.children[:]:
            if not hasattr(child, "_path"):
                continue

            name = child._name_input.text.strip()
            if not name:
                continue

            if self.target_type == "phones":
                phone_id = child._id_input.text.strip() if child._id_input else ""
                if not phone_id:
                    phone_id = f"BULK-{datetime.now().strftime('%H%M%S%f')}"
                img_path = copy_image_to_storage(child._path, get_phone_images_path())
                app.db.add_phone(phone_id=phone_id, name=name, image_path=img_path)
            else:
                img_path = copy_image_to_storage(child._path, get_spare_images_path())
                app.db.add_spare_part(name=name, image_path=img_path)
            count += 1

        self.ids.bulk_status.text = f"Saved {count} items!"
        app.show_toast(f"Saved {count} items!")

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


class BackupScreen(Screen):
    def create_backup(self):
        app = App.get_running_app()
        try:
            backup_dir = get_backup_path()
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"nokia_backup_{timestamp}.zip")

            with zipfile.ZipFile(backup_file, "w", zipfile.ZIP_DEFLATED) as zf:
                # Add database
                db_path = get_db_path()
                if os.path.exists(db_path):
                    zf.write(db_path, "nokia_storage.db")

                # Add all images
                images_dir = get_images_path()
                if os.path.exists(images_dir):
                    for root_dir, dirs, files in os.walk(images_dir):
                        for f in files:
                            full_path = os.path.join(root_dir, f)
                            arc_name = os.path.relpath(full_path, get_app_path())
                            zf.write(full_path, arc_name)

            self.ids.backup_status.text = f"Backup saved to:\n{backup_file}"
            self.ids.backup_status.color = (0.26, 0.63, 0.28, 1)
            app._last_backup_path = backup_file
            app.show_toast("Backup created!")
        except Exception as e:
            self.ids.backup_status.text = f"Error: {str(e)}"
            self.ids.backup_status.color = (0.9, 0.22, 0.21, 1)

    def restore_backup(self):
        app = App.get_running_app()
        app.pick_image_for = ("restore_backup", None)
        app.open_file_chooser(filters=["*.zip"])

    def on_backup_selected(self, path):
        app = App.get_running_app()
        try:
            # Close current database
            app.db.close()

            with zipfile.ZipFile(path, "r") as zf:
                # Extract all to app directory
                zf.extractall(get_app_path())

            # Reopen database
            app.db = NokiaDatabase(get_db_path())

            self.ids.backup_status.text = "Restore complete! Data and images recovered."
            self.ids.backup_status.color = (0.26, 0.63, 0.28, 1)
            app.show_toast("Backup restored!")
        except Exception as e:
            # Reopen database even on failure
            app.db = NokiaDatabase(get_db_path())
            self.ids.backup_status.text = f"Error: {str(e)}"
            self.ids.backup_status.color = (0.9, 0.22, 0.21, 1)

    def share_backup(self):
        app = App.get_running_app()
        backup_path = getattr(app, "_last_backup_path", None)
        if not backup_path or not os.path.exists(backup_path):
            # Create a fresh backup first
            self.create_backup()
            backup_path = getattr(app, "_last_backup_path", None)
            if not backup_path:
                return

        if platform == "android":
            try:
                context = mActivity.getApplicationContext()
                package = context.getPackageName()
                java_file = autoclass("java.io.File")(backup_path)
                uri = FileProvider.getUriForFile(
                    context, f"{package}.fileprovider", java_file
                )
                intent = Intent(Intent.ACTION_SEND)
                intent.setType("application/zip")
                intent.putExtra(Intent.EXTRA_STREAM, cast("android.os.Parcelable", uri))
                intent.putExtra(Intent.EXTRA_SUBJECT, "Nokia Storage Backup")
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                chooser = Intent.createChooser(intent, "Share Backup")
                mActivity.startActivity(chooser)
            except Exception as e:
                self.ids.backup_status.text = f"Share error: {str(e)}"
        else:
            self.ids.backup_status.text = f"Backup at: {backup_path}"
            app.show_toast(f"Backup file: {backup_path}")

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


class SearchAllScreen(Screen):
    initial_query = StringProperty("")

    def on_enter(self):
        if self.initial_query:
            Clock.schedule_once(self._set_initial_query, 0.1)

    def _set_initial_query(self, *args):
        self.ids.search_all_bar.ids.search_input.text = self.initial_query
        self.do_search(self.initial_query)

    def do_search(self, text):
        app = App.get_running_app()
        grid = self.ids.results_list
        grid.clear_widgets()
        default_img = get_default_image()

        if not text.strip():
            grid.add_widget(Label(
                text="Type to search phones and spare parts",
                font_size=sp(14), color=(0.5, 0.5, 0.5, 1),
                size_hint_y=None, height=dp(40),
            ))
            return

        phones, spares = app.db.search_all(text)

        if phones:
            header = Label(
                text=f"Phones ({len(phones)})",
                font_size=sp(15), bold=True,
                color=(0, 0.314, 0.784, 1),
                size_hint_y=None, height=dp(30),
                text_size=(dp(300), None), halign="left",
            )
            grid.add_widget(header)

            for p in phones:
                img = p.get("image_path", "") or default_img
                card = PhoneCard(
                    phone_id=p["id"],
                    phone_name=p["name"],
                    phone_date=p.get("release_date", ""),
                    phone_image=img,
                )
                card.bind(on_release=partial(self._open_phone, p["id"]))
                grid.add_widget(card)

        if spares:
            header = Label(
                text=f"Spare Parts ({len(spares)})",
                font_size=sp(15), bold=True,
                color=(0, 0.314, 0.784, 1),
                size_hint_y=None, height=dp(30),
                text_size=(dp(300), None), halign="left",
            )
            grid.add_widget(header)

            for s in spares:
                img = s.get("image_path", "") or default_img
                card = SpareCard(
                    spare_id=s["id"],
                    spare_name=s["name"],
                    spare_desc=s.get("description", ""),
                    spare_image=img,
                )
                grid.add_widget(card)

        if not phones and not spares:
            grid.add_widget(Label(
                text="No results found",
                font_size=sp(14), color=(0.5, 0.5, 0.5, 1),
                size_hint_y=None, height=dp(40),
            ))

    def _open_phone(self, phone_id, *args):
        app = App.get_running_app()
        detail = app.root.get_screen("phone_detail")
        detail.load_phone(phone_id)
        app.root.transition = SlideTransition(direction="left")
        app.root.current = "phone_detail"

    def go_back(self):
        app = App.get_running_app()
        app.root.transition = SlideTransition(direction="right")
        app.root.current = "main"


# ── Main Application ────────────────────────────────────────────

class NokiaStorageApp(App):
    title = "Nokia Storage"
    db = ObjectProperty(None)
    pick_image_for = None  # ("target", extra_data)
    _file_chooser_popup = None

    def build(self):
        Window.clearcolor = (0.94, 0.96, 1, 1)
        self.db = NokiaDatabase(get_db_path())
        get_default_image()

        if platform == "android":
            self._request_permissions()

        return Builder.load_string(KV)

    def _request_permissions(self):
        if platform == "android":
            perms = [
                Permission.CAMERA,
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE,
            ]
            # Android 13+ media permissions
            try:
                perms.extend([
                    "android.permission.READ_MEDIA_IMAGES",
                    "android.permission.READ_MEDIA_VIDEO",
                ])
            except Exception:
                pass
            request_permissions(perms)

    def show_toast(self, text):
        """Show a brief notification."""
        content = BoxLayout(
            size_hint=(0.8, None), height=dp(50),
            padding=dp(12),
        )
        with content.canvas.before:
            Color(0.2, 0.2, 0.2, 0.9)
            content._bg = RoundedRectangle(
                pos=content.pos, size=content.size, radius=[dp(8)]
            )
        content.bind(pos=lambda w, v: setattr(w._bg, "pos", v))
        content.bind(size=lambda w, v: setattr(w._bg, "size", v))
        content.add_widget(Label(
            text=text, color=(1, 1, 1, 1), font_size=sp(14),
        ))
        popup = ModalView(
            size_hint=(0.8, None), height=dp(50),
            background_color=(0, 0, 0, 0),
            pos_hint={"center_x": 0.5, "y": 0.05},
        )
        popup.add_widget(content)
        popup.open()
        Clock.schedule_once(lambda dt: popup.dismiss(), 2)

    # ── File Chooser ────────────────────────────────────────────

    def open_file_chooser(self, filters=None, multiple=False):
        if platform == "android":
            self._android_file_chooser(filters, multiple)
        else:
            self._desktop_file_chooser(filters, multiple)

    def _desktop_file_chooser(self, filters=None, multiple=False):
        from kivy.uix.filechooser import FileChooserListView

        fc = FileChooserListView(
            filters=filters or ["*.png", "*.jpg", "*.jpeg", "*.bmp"],
            path=os.path.expanduser("~"),
        )
        if multiple:
            fc.multiselect = True

        content = BoxLayout(orientation="vertical", spacing=dp(8))
        content.add_widget(fc)

        btn_row = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))

        cancel = BoxLayout(padding=(dp(8), dp(6)))
        cancel.add_widget(Label(text="Cancel", font_size=sp(14), color=(0.4, 0.4, 0.4, 1)))

        select = BoxLayout(padding=(dp(8), dp(6)))
        with select.canvas.before:
            Color(0, 0.314, 0.784, 1)
            select._bg = RoundedRectangle(pos=select.pos, size=select.size, radius=[dp(8)])
        select.bind(pos=lambda w, v: setattr(w._bg, "pos", v))
        select.bind(size=lambda w, v: setattr(w._bg, "size", v))
        select.add_widget(Label(text="Select", font_size=sp(14), color=(1, 1, 1, 1), bold=True))

        btn_row.add_widget(cancel)
        btn_row.add_widget(select)
        content.add_widget(btn_row)

        popup = Popup(
            title="Select File", content=content,
            size_hint=(0.95, 0.85),
        )

        cancel.bind(on_touch_down=lambda w, t: popup.dismiss() if w.collide_point(*t.pos) else None)
        select.bind(on_touch_down=lambda w, t: (
            self._on_file_selected(fc.selection, popup) if w.collide_point(*t.pos) else None
        ))
        popup.open()

    def _android_file_chooser(self, filters=None, multiple=False):
        try:
            from plyer import filechooser
            if multiple:
                filechooser.open_file(
                    on_selection=self._on_android_selection,
                    multiple=True,
                    filters=filters or ["image/*"],
                )
            else:
                ext_filters = filters
                if not ext_filters or ext_filters == ["*.png", "*.jpg", "*.jpeg", "*.bmp"]:
                    ext_filters = ["image/*"]
                filechooser.open_file(
                    on_selection=self._on_android_selection,
                    filters=ext_filters,
                )
        except Exception:
            # Fallback to intent-based chooser
            self._android_intent_chooser(filters, multiple)

    def _android_intent_chooser(self, filters=None, multiple=False):
        try:
            intent = Intent(Intent.ACTION_GET_CONTENT)
            if filters and "*.zip" in filters:
                intent.setType("application/zip")
            elif filters and ("*.xlsx" in filters or "*.xls" in filters):
                intent.setType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                intent.setType("image/*")
            if multiple:
                intent.putExtra(Intent.EXTRA_ALLOW_MULTIPLE, True)
            mActivity.startActivityForResult(intent, 1001)
        except Exception as e:
            self.show_toast(f"File picker error: {str(e)}")

    def _on_android_selection(self, selection):
        if selection:
            self._on_file_selected(selection)

    def _on_file_selected(self, selection, popup=None):
        if popup:
            popup.dismiss()

        if not selection:
            return

        target = self.pick_image_for
        if not target:
            return

        target_type, target_data = target

        if target_type == "add_phone_screen":
            screen = self.root.get_screen("add_phone")
            screen.on_image_selected(selection[0])

        elif target_type == "add_spare_screen":
            screen = self.root.get_screen("add_spare")
            screen.on_image_selected(selection[0])

        elif target_type == "phone":
            # Update phone image directly
            phone_id = target_data
            img_path = copy_image_to_storage(selection[0], get_phone_images_path())
            if img_path:
                self.db.update_phone(phone_id, image_path=img_path)
                detail = self.root.get_screen("phone_detail")
                detail.image_source = img_path
                self.show_toast("Image updated!")

        elif target_type == "import_excel":
            screen = self.root.get_screen("import_excel")
            screen.on_file_selected(selection[0])

        elif target_type == "restore_backup":
            screen = self.root.get_screen("backup")
            screen.on_backup_selected(selection[0])

        elif target_type == "bulk_images":
            screen = self.root.get_screen("bulk_images")
            screen.on_images_selected(selection)

        self.pick_image_for = None

    # ── Camera ──────────────────────────────────────────────────

    def take_camera_photo(self):
        if platform == "android":
            try:
                intent = Intent(MediaStore.ACTION_IMAGE_CAPTURE)
                # Save to temp file
                temp_dir = os.path.join(get_app_path(), "temp")
                os.makedirs(temp_dir, exist_ok=True)
                temp_file = os.path.join(
                    temp_dir, f"cam_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                )
                context = mActivity.getApplicationContext()
                package = context.getPackageName()
                java_file = autoclass("java.io.File")(temp_file)
                uri = FileProvider.getUriForFile(
                    context, f"{package}.fileprovider", java_file
                )
                intent.putExtra(MediaStore.EXTRA_OUTPUT, cast("android.os.Parcelable", uri))
                mActivity.startActivityForResult(intent, 1002)
                self._camera_temp_file = temp_file
            except Exception as e:
                self.show_toast(f"Camera error: {str(e)}")
        else:
            self.show_toast("Camera only available on Android")

    # ── Excel Export ────────────────────────────────────────────

    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()

            # Sheet 1: Phones
            ws_phones = wb.active
            ws_phones.title = "Nokia Phones"
            headers = ["ID", "Name", "Release Date", "Appearance Condition",
                        "Working Condition", "Remarks"]
            header_fill = PatternFill(start_color="0050C8", end_color="0050C8",
                                       fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, h in enumerate(headers, 1):
                cell = ws_phones.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            phones = self.db.export_phones()
            for i, p in enumerate(phones, 2):
                ws_phones.cell(row=i, column=1, value=p["id"])
                ws_phones.cell(row=i, column=2, value=p["name"])
                ws_phones.cell(row=i, column=3, value=p.get("release_date", ""))
                ws_phones.cell(row=i, column=4, value=p.get("appearance_condition", ""))
                ws_phones.cell(row=i, column=5, value=p.get("working_condition", ""))
                ws_phones.cell(row=i, column=6, value=p.get("remarks", ""))

            # Auto-width
            for col in range(1, 7):
                ws_phones.column_dimensions[chr(64 + col)].width = 18

            # Sheet 2: Spare Parts
            ws_spares = wb.create_sheet("Spare Parts")
            spare_headers = ["ID", "Name", "Phone ID", "Description"]
            for col, h in enumerate(spare_headers, 1):
                cell = ws_spares.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            spares = self.db.export_spare_parts()
            for i, s in enumerate(spares, 2):
                ws_spares.cell(row=i, column=1, value=s["id"])
                ws_spares.cell(row=i, column=2, value=s["name"])
                ws_spares.cell(row=i, column=3, value=s.get("phone_id", ""))
                ws_spares.cell(row=i, column=4, value=s.get("description", ""))

            for col in range(1, 5):
                ws_spares.column_dimensions[chr(64 + col)].width = 20

            # Save
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(export_dir, f"nokia_export_{timestamp}.xlsx")
            wb.save(filepath)

            self.show_toast(f"Exported to {filepath}")

            # Share on Android
            if platform == "android":
                self._share_file(filepath, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except ImportError:
            self.show_toast("openpyxl not available for export")
        except Exception as e:
            self.show_toast(f"Export error: {str(e)}")

    def _share_file(self, filepath, mime_type):
        if platform == "android":
            try:
                context = mActivity.getApplicationContext()
                package = context.getPackageName()
                java_file = autoclass("java.io.File")(filepath)
                uri = FileProvider.getUriForFile(
                    context, f"{package}.fileprovider", java_file
                )
                intent = Intent(Intent.ACTION_SEND)
                intent.setType(mime_type)
                intent.putExtra(Intent.EXTRA_STREAM, cast("android.os.Parcelable", uri))
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                chooser = Intent.createChooser(intent, "Share")
                mActivity.startActivity(chooser)
            except Exception:
                pass

    def on_stop(self):
        if self.db:
            self.db.close()


if __name__ == "__main__":
    NokiaStorageApp().run()
